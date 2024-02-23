import pandas as pd
import requests as rq 
import pprint
import sqlite3
from localidades import nacional
import ssl
from Google import Create_Service
from googleapiclient.http import MediaFileUpload
import openpyxl
from ajustar_planilha import ajustar_colunas, ajustar_bordas

tabela5917 = 5917
tabela4093 = 4093
tabela6398 = 6398
tabela5918 = 5918
tabela4094 = 4094
tabela6388 = 6388
tabela6499 = 6399
tabela6403 = 6403
tabela6402 = 6402
tabela5919 = 5919
tabela4095 = 4095

api_populacao_sexo = f'https://servicodados.ibge.gov.br/api/v3/agregados/{tabela5917}/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202202|202203|202204|202301|202302|202303/variaveis/606?{nacional}&classificacao=2[4,5]'
api_trab_sexo = f'https://servicodados.ibge.gov.br/api/v3/agregados/{tabela4093}/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303/variaveis/1641|4088|4090|4092|4094?{nacional}&classificacao=2[4,5]'
api_forca_sexo = f'https://servicodados.ibge.gov.br/api/v3/agregados/{tabela6388}/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/8344|8346?localidades=N1[all]&classificacao=2[4,5]'

api_populacao_idade= f'https://servicodados.ibge.gov.br/api/v3/agregados/{tabela5918}/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/606?{nacional}&classificacao=58[40288,114535,100052,108875,99127,3302]'
api_anos_idade = f' https://servicodados.ibge.gov.br/api/v3/agregados/{tabela4094}/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/1641|4088|4090|4092|4094?{nacional}&classificacao=58[114535,100052,108875,99127,3302]'
api_trab_idade = f'https://servicodados.ibge.gov.br/api/v3/agregados/{tabela6499}/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202202|202203|202204|202301|202302|202303|202304/variaveis/8344|8346?localidades=N1[all]&classificacao=58[114535,100052,108875,99127,3302]'

api_populacao_raca= f'https://servicodados.ibge.gov.br/api/v3/agregados/{tabela6403}/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/606?{nacional}&classificacao=86[2776,2777,2779]'
api_trab_raca = f'https://servicodados.ibge.gov.br/api/v3/agregados/{tabela6402}/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/4088|4090|4092|4094?{nacional}&classificacao=86[2776,2777,2779]'

api_populacao_alfab= f'https://servicodados.ibge.gov.br/api/v3/agregados/{tabela5919}/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/606?{nacional}&classificacao=1568[120706,11779,11628,11629,11630,11631,11632,11626]'
api_trab_alfab = f'https://servicodados.ibge.gov.br/api/v3/agregados/{tabela4095}/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/1641|4088|4090|4092|4094?{nacional}&classificacao=1568[120706,11779,11628,11629,11630,11631,11632,11626]'
    


class TLSAdapter(rq.adapters.HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        ctx = ssl.create_default_context()
        ctx.set_ciphers("DEFAULT@SECLEVEL=1")
        ctx.options |= 0x4   # OP_LEGACY_SERVER_CONNECT
        kwargs["ssl_context"] = ctx
        return super(TLSAdapter, self).init_poolmanager(*args, **kwargs)

def requisitando_dados(api):
    with rq.session() as s:
        s.mount("https://", TLSAdapter())
        dados_brutos_api = s.get(api, verify=True)

    # Verificação se a solicitação foi bem-sucedida antes de continuar
    if dados_brutos_api.status_code != 200:
        raise Exception(f"A solicitação à API falhou com o código de status: {dados_brutos_api.status_code}")

    # Verificação se a resposta pode ser convertida para JSON
    try:
        dados_brutos = dados_brutos_api.json()
    except Exception as e:
        raise Exception(f"Erro ao analisar a resposta JSON da API: {str(e)}")

    return dados_brutos

def extrair_dados(api, tabela_id):
    dados_brutos = requisitando_dados(api)
    lista_tabela = [tabela5917, tabela5918, tabela4094, tabela6388, tabela6499, tabela6403, tabela6402, tabela5919, tabela4095]
    if dados_brutos:
        for tabelas in lista_tabela:
            if tabela_id == tabelas:
                variavel = dados_brutos[0]
                return variavel
    else:
        pass
    
def tratando_dadosPopu(variavel1209):
    dados_limpos_1209 = []

    
    id_tabela = variavel1209['id']
    variavel = variavel1209['variavel']
    unidade = variavel1209['unidade']
    dados = variavel1209['resultados']
    
    for ii in dados:
        dados_produto = ii['classificacoes']
        dados_producao = ii['series']

        for iii in dados_produto:
            dados_id_produto = iii['categoria']

            for id_produto, nome_produto in dados_id_produto.items():
                

                for iv in dados_producao:
                    id = iv['localidade']['id']
                    local = iv['localidade']['nome']
                    dados_ano_producao = iv['serie']
                    

                    for ano, producao in dados_ano_producao.items():
                        producao = producao.replace('-', '0').replace('...', '0')
                        
                        
                        dict = {
                            'id': id,
                            'local': local,
                            'id_produto': id_produto,
                            'Categoria': nome_produto,
                            variavel: producao,
                            'unidade': unidade,
                            'ano': f'01/01/{ano}'
                        }

                        dados_limpos_1209.append(dict)
    return dados_limpos_1209     
    
def tratando_dados(variavel):
    dados_limpos = []

    
    id_tabela = variavel['id']
    variavele = variavel['variavel']
    unidade = variavel['unidade']
    dados = variavel['resultados']
    
    for ii in dados:
        dados_produto = ii['classificacoes']
        dados_producao = ii['series']

        for iii in dados_produto:
            dados_id_produto = iii['categoria']

            for id_produto, nome_produto in dados_id_produto.items():

                for iv in dados_producao:
                    id = iv['localidade']['id']
                    local = iv['localidade']['nome']
                    dados_ano_producao = iv['serie']
                    

                    for ano, producao in dados_ano_producao.items():
                        producao = producao.replace('-', '0').replace('...', '0')
                        
                        partes = ano.split("/")
                        ano_sem_trimestre = int(partes[0][:4])
                        trimestre = int(partes[0][4:6])
                        
                        dict = {
                            'id': id,
                            'local': local,
                            'id_produto': id_produto,
                            'Categoria': nome_produto,
                            variavele: producao,
                            'unidade': unidade,
                            'ano': f'01/01/{ano_sem_trimestre}',
                            'Trimestre': trimestre
                        }

                        dados_limpos.append(dict)
    return dados_limpos 

def tratando_dadosTrab(lista_var):
    dados_limpos = []
    
    for i in lista_var:
        id_tabela = i['id']
        variavele = i['variavel']
        unidade = i['unidade']
        dados = i['resultados']
        
        for ii in dados:
            dados_produto = ii['classificacoes']
            dados_producao = ii['series']

            for iii in dados_produto:
                dados_id_produto = iii['categoria']

                for id_produto, nome_produto in dados_id_produto.items():

                    for iv in dados_producao:
                        id = iv['localidade']['id']
                        local = iv['localidade']['nome']
                        dados_ano_producao = iv['serie']
                        

                        for ano, producao in dados_ano_producao.items():
                            producao = producao.replace('-', '0').replace('...', '0')
                            
                            partes = ano.split("/")
                            ano_sem_trimestre = int(partes[0][:4])
                            trimestre = int(partes[0][4:6])
                            
                            dict = {
                                'id': id,
                                'local': local,
                                'id_produto': id_produto,
                                'Categoria': nome_produto,
                                variavele: producao,
                                'unidade': unidade,
                                'ano': f'01/01/{ano_sem_trimestre}',
                                'Trimestre': trimestre
                            }

                            '''if id_tabela == '8331':
                                dados_limpos_8331.append(dict)
                            elif id_tabela == '216':
                                dados_limpos_216.append(dict)
                            elif id_tabela == '214':
                                dados_limpos_214.append(dict)
                            elif id_tabela == '112':
                                dados_limpos_112.append(dict)'''

    return dados_limpos

def executando_funcoes():
    variavel_pop_sexo = extrair_dados(api_populacao_sexo, tabela5917)
    variavel_trab_sexo = extrair_dados(api_trab_sexo, tabela4093)
    
    variavel_pop_idade = extrair_dados(api_populacao_idade, tabela5918)
    varaivel_anos = extrair_dados(api_anos_idade, tabela4094)
    variavel_trab = extrair_dados(api_trab_idade, tabela6388)
    
    variavel_pop_raca = extrair_dados(api_populacao_raca, tabela6403)
    variavel_trab_raca = extrair_dados(api_trab_raca, tabela6402)
    
    variavel_pop_alfab = extrair_dados(api_populacao_alfab, tabela5919)
    variavel_trab_alfab = extrair_dados(api_trab_alfab, tabela4095)
    
    dados_limpos_pop_sexo = tratando_dados(variavel_pop_sexo)
    dados_limpos_trab_sexo = tratando_dadosTrab(variavel_trab_sexo)
    
    dados_limpos_pop_idade = tratando_dadosPopu(variavel_pop_idade)
    dados_limpos_anos_idade = tratando_dadosTrab(5)
    dados_limpos_trab_idade = tratando_dadosTrab(2)
    
    dados_limpos_pop_raca = tratando_dadosPopu(variavel_pop_raca)
    dados_limpos_trab_raca= tratando_dadosTrab(4)
    
    dados_limpos_pop_alfab = tratando_dadosPopu(variavel_pop_alfab)
    dados_limpos_trab_alfab = tratando_dadosTrab(5)
    
    return dados_limpos_pop_sexo, dados_limpos_trab_sexo, dados_limpos_pop_idade, dados_limpos_anos_idade, dados_limpos_trab_idade, dados_limpos_pop_raca, dados_limpos_trab_raca, dados_limpos_pop_alfab,dados_limpos_trab_alfab
    
def gerando_dataframe(dados_limpos_pop_sexo, dados_limpos_trab_sexo, dados_limpos_pop_idade, dados_limpos_anos_idade, dados_limpos_trab_idade, dados_limpos_pop_raca, dados_limpos_trab_raca, dados_limpos_pop_alfab,dados_limpos_trab_alfab):
    dfpop_sexo = pd.DataFrame(dados_limpos_pop_sexo)
    dftrab_sexo = pd.DataFrame(dados_limpos_trab_sexo)
    dfpop_idade = pd.DataFrame(dados_limpos_pop_idade)
    dfanos_idade = pd.DataFrame(dados_limpos_anos_idade)
    dftrab_idade = pd.DataFrame(dados_limpos_trab_idade)
    dfpop_raca = pd.DataFrame(dados_limpos_pop_raca)
    dftrab_raca = pd.DataFrame(dados_limpos_trab_raca)
    dfpop_alfab = pd.DataFrame(dados_limpos_pop_alfab)
    dftrab_alfab = pd.DataFrame(dados_limpos_trab_alfab)
    
    
    return dfpop_sexo, dftrab_sexo, dfpop_idade, dfanos_idade, dftrab_idade, dfpop_raca, dftrab_raca, dfpop_alfab, dftrab_alfab

pp = pprint.PrettyPrinter(indent=4)
dados_limpos_pop_sexo, dados_limpos_trab_sexo, dados_limpos_pop_idade, dados_limpos_anos_idade, dados_limpos_trab_idade, dados_limpos_pop_raca, dados_limpos_trab_raca, dados_limpos_pop_alfab,dados_limpos_trab_alfab = executando_funcoes()
dfpop_sexo, dftrab_sexo, dfpop_idade, dfanos_idade, dftrab_idade, dfpop_raca, dftrab_raca, dfpop_alfab, dftrab_alfab = gerando_dataframe(dados_limpos_pop_sexo, dados_limpos_trab_sexo, dados_limpos_pop_idade, dados_limpos_anos_idade, dados_limpos_trab_idade, dados_limpos_pop_raca, dados_limpos_trab_raca, dados_limpos_pop_alfab,dados_limpos_trab_alfab)
print(dftrab_sexo)


dfpop_sexo.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\População SEXO NACIONAL.xlsx', index=False)
# dftrab_sexo.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Força de trabalho SEXO NACIONAL.xlsx', index=False)
dfpop_idade.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Populaçao IDADE NACIONAL.xlsx', index=False)
dfanos_idade.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Anos IDADE NACIONAL.xlsx', index=False)
dftrab_idade.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Trabalho IDADE NACIONAL.xlsx', index=False)
dfpop_raca.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\População RACA NACIONAL.xlsx', index=False)
dftrab_raca.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Trabalho RACA NACIONAL.xlsx', index=False)
dfpop_alfab.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\População Grau de instrução NACIONAL.xlsx', index=False)
# dftrab_alfab.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Trabalha Grau de instrução NACIONAL.xlsx', index=False)
'''

planilha_principal = openpyxl.Workbook()

wb_1209 = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\População NACIONAL.xlsx')
wb_5918 = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Idade NACIONAL.xlsx')
wb_6463 = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Força de trabalho NACIONAL.xlsx')
wb_6482 = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Subutilização NACIONAL.xlsx')

aba_1209 = planilha_principal.create_sheet("População Total")
aba_5918 = planilha_principal.create_sheet("Pessoas aptam a trabalhar")
aba_6463 = planilha_principal.create_sheet("Relação de força de trabalho")
aba_6482 = planilha_principal.create_sheet("Subutilização")


for linha in wb_1209.active.iter_rows(values_only=True):
    aba_1209.append(linha)

for linha in wb_5918.active.iter_rows(values_only=True):
    aba_5918.append(linha)
    
for linha in wb_6463.active.iter_rows(values_only=True):
    aba_6463.append(linha)
    
for linha in wb_6482.active.iter_rows(values_only=True):
    aba_6482.append(linha)
    
for aba in planilha_principal.sheetnames:
    if aba not in ["População Total", "Pessoas aptam a trabalhar", "Relação de força de trabalho", "Subutilização"]:
        del planilha_principal[aba]
        
ajustar_bordas(planilha_principal)

lista_aba = [aba_1209, aba_5918, aba_6463, aba_6482]
for abas in lista_aba:
    ajustar_colunas(abas)
    
planilha_principal.save("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas Tratadas\\PNAD NACIONAL.xlsx")   
'''