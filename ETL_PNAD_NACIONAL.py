import pandas as pd
import requests as rq 
import pprint
from localidades import nacional
import ssl
from Google import Create_Service
from googleapiclient.http import MediaFileUpload
import openpyxl
from ajustar_planilha import ajustar_colunas, ajustar_bordas

tabela1209 = 1209
tabela5918 = 5918
tabela6463 = 6463
tabela6482 = 6482

api_populacao = f'https://servicodados.ibge.gov.br/api/v3/agregados/1209/periodos/2022/variaveis/606?localidades=N1[all]&classificacao=58[0]'

api_anos = f'https://servicodados.ibge.gov.br/api/v3/agregados/{tabela5918}/periodos/201201%7C201202%7C201203%7C201204%7C201301%7C201302%7C201303%7C201304%7C201401%7C201402%7C201403%7C201404%7C201501%7C201502%7C201503%7C201504%7C201601%7C201602%7C201603%7C201604%7C201701%7C201702%7C201703%7C201704%7C201801%7C201802%7C201803%7C201804%7C201901%7C201902%7C201903%7C201904%7C202001%7C202002%7C202003%7C202004%7C202101%7C202102%7C202103%7C202104%7C202201%7C202202%7C202203%7C202204%7C202301%7C202302%7C202303%7C202304/variaveis/606?{nacional}&classificacao=58[40288,114535,100052,108875,99127,3302]'

api_trab = f'https://servicodados.ibge.gov.br/api/v3/agregados/6463/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/1641?localidades=N1[all]&classificacao=629[32386,32387,32446,32447]'
    
api_potencial = f'https://servicodados.ibge.gov.br/api/v3/agregados/6482/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/1641?localidades=N1[all]&classificacao=604[31751,31752,46254]'

class TLSAdapter(rq.adapters.HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        ctx = ssl.create_default_context()
        ctx.set_ciphers("DEFAULT@SECLEVEL=1")
        ctx.options |= 0x4   # OP_LEGACY_SERVER_CONNECT
        kwargs["ssl_context"] = ctx
        return super(TLSAdapter, self).init_poolmanager(*args, **kwargs)

def requisitando_dados(api):
    with rq.session() as s:
        s.mount("https://", TLSAdapter())
        dados_brutos_api = s.get(api, verify=True)
    # Verificação se a solicitação foi bem-sucedida antes de continuar
    if dados_brutos_api.status_code != 200:
        raise Exception(f"A solicitação à API falhou com o código de status: {dados_brutos_api.status_code}")

    # Verificação se a resposta pode ser convertida para JSON
    try:
        dados_brutos = dados_brutos_api.json()
    except Exception as e:
        raise Exception(f"Erro ao analisar a resposta JSON da API: {str(e)}")

    return dados_brutos

def extrair_dados(api, tabela_id):
    dados_brutos = requisitando_dados(api)

    if dados_brutos:
        if tabela_id == tabela1209:
            variavel606 = dados_brutos[0]
            return variavel606
        elif tabela_id == tabela5918:
            variavel606 = dados_brutos[0]
            return variavel606
        elif tabela_id == tabela6463:
            variavel1641 = dados_brutos[0]
            return variavel1641
        elif tabela_id == tabela6482:
            variavel1641 = dados_brutos[0]
            return variavel1641
    else:
        pass
    
def tratando_dados1209(variavel1209):
    dados_limpos_1209 = []

    id_tabela = variavel1209['id']
    variavel = variavel1209['variavel']
    unidade = variavel1209['unidade']
    dados = variavel1209['resultados']
    
    for ii in dados:
        dados_produto = ii['classificacoes']
        dados_producao = ii['series']

        for iii in dados_produto:
            dados_id_produto = iii['categoria']

            for id_produto, nome_produto in dados_id_produto.items():
                #nome_produto = 'Outros produtos ' + nome_produto if not nome_produto else nome_produto

                for iv in dados_producao:
                    id = iv['localidade']['id']
                    local = iv['localidade']['nome']
                    dados_ano_producao = iv['serie']
                
                    for ano, producao in dados_ano_producao.items():
                        producao = producao.replace('-', '0').replace('...', '0')
                        
                        dict = {
                            'id': id,
                            'local': local,
                            #'id_produto': id_produto,
                            'Categoria': nome_produto,
                            variavel: producao,
                            'unidade': unidade,
                            'ano': f'01/01/{ano}'
                        }

                        dados_limpos_1209.append(dict)
    return dados_limpos_1209     
    
def tratando_dados(variavel):
    dados_limpos = []

    id_tabela = variavel['id']
    variavele = variavel['variavel']
    unidade = variavel['unidade']
    dados = variavel['resultados']
    
    for ii in dados:
        dados_produto = ii['classificacoes']
        dados_producao = ii['series']

        for iii in dados_produto:
            dados_id_produto = iii['categoria']

            for id_produto, nome_produto in dados_id_produto.items():

                for iv in dados_producao:
                    id = iv['localidade']['id']
                    local = iv['localidade']['nome']
                    dados_ano_producao = iv['serie']
                    

                    for ano, producao in dados_ano_producao.items():
                        producao = producao.replace('-', '0').replace('...', '0')
                        
                        partes = ano.split("/")
                        ano_sem_trimestre = int(partes[0][:4])
                        trimestre = int(partes[0][4:6])
                        
                        dict = {
                            'id': id,
                            'local': local,
                            #'id_produto': id_produto,
                            'Categoria': nome_produto,
                            variavele: producao,
                            'unidade': unidade,
                            'ano': f'01/01/{ano_sem_trimestre}',
                            'Trimestre': trimestre
                        }

                        dados_limpos.append(dict)
    return dados_limpos 

def executando_funcoes():
    variavel1209 = extrair_dados(api_populacao, tabela1209)
    variavel5918 = extrair_dados(api_anos, tabela5918)
    variavel6463 = extrair_dados(api_trab, tabela6463)
    varaivel6482 = extrair_dados(api_potencial, tabela6482)
    dados_limpos_1209 = tratando_dados1209(variavel1209)
    dados_limpos_5918 = tratando_dados(variavel5918)
    dados_limpos_6463 = tratando_dados(variavel6463)
    dados_limpos_6482 = tratando_dados(varaivel6482)
    
    return dados_limpos_1209, dados_limpos_5918, dados_limpos_6463, dados_limpos_6482
    
def gerando_dataframe(dados_limpos_1209, dados_limpos_5918, dados_limpos_6463, dados_limpos_6482):
    df1209 = pd.DataFrame(dados_limpos_1209)
    df5918 = pd.DataFrame(dados_limpos_5918)
    df6463 = pd.DataFrame(dados_limpos_6463)
    df6482 = pd.DataFrame(dados_limpos_6482)
    
    # df1209['População'] = df1209['População'].astype(float)
    # df5918['População'] = df5918['População'].astype(float)
    # df6463['Pessoas de 14 anos ou mais de idade'] = df6463['Pessoas de 14 anos ou mais de idade'].astype(float)
    # df6482['Pessoas de 14 anos ou mais de idade'] = df6463['Pessoas de 14 anos ou mais de idade'].astype(float)
    
    df5918[['0 a 13 anos', '14 a 17 anos', '18 a 24 anos', '25 a 39 anos', '40 a 59 anos', '60 anos ou mais']] = None
    linhas_5918_0, linhas_5918_14, linhas_5918_18, linhas_5918_25, linhas_5918_40, linhas_5918_60 = slice(0, 47), slice(48, 95), slice(96, 143), slice(144, 191), slice(192, 239), slice(240, 288)
    lista_5918 = [linhas_5918_0, linhas_5918_14, linhas_5918_18, linhas_5918_25, linhas_5918_40, linhas_5918_60]
    lista_mover5918 = ['0 a 13 anos', '14 a 17 anos', '18 a 24 anos', '25 a 39 anos', '40 a 59 anos', '60 anos ou mais']
    for linhas, mover in zip(lista_5918, lista_mover5918):
        df5918.loc[linhas, mover] = df5918.loc[linhas, 'População']
    
    colunas_5918m = ['14 a 17 anos', '18 a 24 anos', '25 a 39 anos', '40 a 59 anos', '60 anos ou mais']
    source_range = [(49, 97), (97, 144), (145, 192), (193, 240), (241, 288)]
    destination_range = (1, 49) 
    for source_range, column_to_move in zip(source_range, colunas_5918m):
            source_values = df5918.loc[source_range[0]-1:source_range[1]-1, column_to_move]
            destination_range_end = destination_range[0] + (source_range[1] - source_range[0])
            df5918.loc[destination_range[0]-1:destination_range_end-1, column_to_move] = source_values.values
    
    del df5918['Categoria']
    del df5918['População']
    for i in range(48, 288):
        df5918.drop(index=i, inplace=True) 
    
    df6463[['Força de trabalho', 'Ocupado', 'Desocupado', 'Fora da Força de trabalho']] = None
    linhas_6463_F, linhas_6463_O, linhas_6463_DO, linhas_6463_FF = slice(0, 47), slice(48, 95), slice(96, 143), slice(144, 192)
    df6463.loc[linhas_6463_F, 'Força de trabalho'] = df6463.loc[linhas_6463_F, 'Pessoas de 14 anos ou mais de idade']
    df6463.loc[linhas_6463_O, 'Ocupado'] = df6463.loc[linhas_6463_O, 'Pessoas de 14 anos ou mais de idade']
    df6463.loc[linhas_6463_DO, 'Desocupado'] = df6463.loc[linhas_6463_DO, 'Pessoas de 14 anos ou mais de idade']
    df6463.loc[linhas_6463_FF, 'Fora da Força de trabalho'] = df6463.loc[linhas_6463_FF, 'Pessoas de 14 anos ou mais de idade']
    source_range = [(49, 97), (97, 144), (145, 192)]
    lista_colunasmudar = ['Ocupado', 'Desocupado', 'Fora da Força de trabalho']
    destination_range = (1, 49) 
    for source_range, column_to_move in zip(source_range, lista_colunasmudar):
            source_values = df6463.loc[source_range[0]-1:source_range[1]-1, column_to_move]
            destination_range_end = destination_range[0] + (source_range[1] - source_range[0])
            df6463.loc[destination_range[0]-1:destination_range_end-1, column_to_move] = source_values.values
    for i in range(48, 192):
        df6463.drop(index=i, inplace=True) 
    del df6463['Categoria']
    del df6463['Pessoas de 14 anos ou mais de idade']

    df6482[['Subocupado por insuficiência de horas trabalhadas', 'Força de trabalho potencial', 'Desalentado']] = None
    linhas_6482_H, linhas_6482_FF, linhas_6482_DE = slice(0, 47), slice(48, 95), slice(96, 143)
    df6482['Subocupado por insuficiência de horas trabalhadas'] = df6482.loc[linhas_6482_H, 'Pessoas de 14 anos ou mais de idade']
    df6482['Força de trabalho potencial'] = df6482.loc[linhas_6482_FF, 'Pessoas de 14 anos ou mais de idade']
    df6482['Desalentado'] = df6482.loc[linhas_6482_DE, 'Pessoas de 14 anos ou mais de idade']
    source_range2 = [(49, 97), (97, 144), (145, 192)]
    lista_muda6482 = ['Força de trabalho potencial', 'Desalentado']
    destination_range = (1, 49) 
    for source_range, column_to_move in zip(source_range2, lista_muda6482):
            source_values = df6482.loc[source_range[0]-1:source_range[1]-1, column_to_move]
            destination_range_end = destination_range[0] + (source_range[1] - source_range[0])
            df6482.loc[destination_range[0]-1:destination_range_end-1, column_to_move] = source_values.values
            
    for i in range(48, 144):
        df6482.drop(index=i, inplace=True) 
    del df6482['Categoria']
    del df6482['Pessoas de 14 anos ou mais de idade']
    
    return df1209, df5918, df6463, df6482

pp = pprint.PrettyPrinter(indent=4)
dados_limpos_1209, dados_limpos_5918, dados_limpos_6463, dados_limpos_6482 = executando_funcoes()
dataframe1209, dataframe5918, dataframe6463, dataframe6482 = gerando_dataframe(dados_limpos_1209, dados_limpos_5918, dados_limpos_6463, dados_limpos_6482)

dftrab = pd.merge(dataframe6463, dataframe6482, on=['id', 'local','unidade', 'ano', 'Trimestre'], how='inner')
#dftrab = pd.merge(dftrab, dataframe5918, on=['id', 'local','unidade', 'ano', 'Trimestre'], how='inner')
# print(dftrab)
dftrab.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\TRABALHO NACIONAL.xlsx', index=False)
dataframe1209.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\População NACIONAL.xlsx', index=False)
dataframe5918.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\IDADE TRABALHO NACIONAL.xlsx', index=False)

#ESTETICA/PLANILHAS
planilha_principal = openpyxl.Workbook()

wb_1209 = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\População NACIONAL.xlsx')
wb_trab = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\TRABALHO NACIONAL.xlsx')
wb_idade = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\IDADE TRABALHO NACIONAL.xlsx')

aba_1209 = planilha_principal.create_sheet("População Total")
aba_trab = planilha_principal.create_sheet("Pessoas aptam a trabalhar")
aba_idade = planilha_principal.create_sheet("IDADE")

for linha in wb_1209.active.iter_rows(values_only=True):
    aba_1209.append(linha)

for linha in wb_trab.active.iter_rows(values_only=True):
    aba_trab.append(linha)
    
for linha in wb_idade.active.iter_rows(values_only=True):
    aba_idade.append(linha)
    
for aba in planilha_principal.sheetnames:
    if aba not in ["População Total", "Pessoas aptam a trabalhar", "IDADE"]:
        del planilha_principal[aba]
        
ajustar_bordas(planilha_principal)

lista_aba = [aba_1209, aba_trab, aba_idade]
for abas in lista_aba:
    ajustar_colunas(abas)
    
planilha_principal.save("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas Tratadas\\PNAD NACIONAL.xlsx")   